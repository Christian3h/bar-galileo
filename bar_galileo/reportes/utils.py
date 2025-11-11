from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
from django.db.models import Sum, Count, Avg, Q, F
from decimal import Decimal
import csv
import io
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def generar_csv_reporte(reporte, datos):
    """Generar reporte en formato CSV con datos detallados"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"reporte_{reporte.tipo}_{reporte.fecha_inicio}_{reporte.fecha_fin}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # BOM para UTF-8
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Encabezados del reporte
    writer.writerow(['BAR GALILEO - REPORTE DE ' + reporte.get_tipo_display().upper()])
    writer.writerow([])
    writer.writerow(['Nombre:', reporte.nombre])
    writer.writerow(['Tipo:', reporte.get_tipo_display()])
    writer.writerow(['Periodo:', f"{reporte.fecha_inicio.strftime('%d/%m/%Y')} - {reporte.fecha_fin.strftime('%d/%m/%Y')}"])
    writer.writerow(['Duración:', f"{reporte.duracion_dias} días"])
    writer.writerow(['Creado por:', str(reporte.creado_por)])
    writer.writerow(['Fecha de generación:', reporte.ultima_generacion.strftime('%d/%m/%Y %H:%M') if reporte.ultima_generacion else 'N/A'])
    writer.writerow([])
    
    # Resumen
    if 'resumen' in datos:
        writer.writerow(['=== RESUMEN ==='])
        for key, value in datos['resumen'].items():
            writer.writerow([key, value])
        writer.writerow([])
    
    # Datos detallados
    if 'detalles' in datos and datos['detalles']:
        writer.writerow(['=== DETALLES ==='])
        
        # Encabezados de la tabla
        if datos['detalles']:
            headers = list(datos['detalles'][0].keys())
            writer.writerow(headers)
            
            # Filas de datos
            for item in datos['detalles']:
                writer.writerow([item.get(h, '') for h in headers])
        writer.writerow([])
    
    # Totales
    if 'totales' in datos:
        writer.writerow(['=== TOTALES ==='])
        for key, value in datos['totales'].items():
            writer.writerow([key, value])
    
    return response


def generar_excel_reporte(reporte, datos):
    """Generar reporte en formato Excel (XLSX) con formato profesional"""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl no está disponible para exportar a Excel")
    
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = reporte.get_tipo_display()[:31]  # Límite de Excel

    # Estilos mejorados
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill(start_color="A68932", end_color="A68932", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    info_label_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Título principal
    ws[f'A{row}'] = f'BAR GALILEO - REPORTE DE {reporte.get_tipo_display().upper()}'
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = title_fill
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f'A{row}:F{row}')
    ws.row_dimensions[row].height = 25
    row += 2
    
    # Información del reporte
    ws[f'A{row}'] = 'Nombre:'
    ws[f'A{row}'].font = info_label_font
    ws[f'B{row}'] = reporte.nombre
    row += 1
    
    ws[f'A{row}'] = 'Tipo:'
    ws[f'A{row}'].font = info_label_font
    ws[f'B{row}'] = reporte.get_tipo_display()
    row += 1
    
    ws[f'A{row}'] = 'Periodo:'
    ws[f'A{row}'].font = info_label_font
    ws[f'B{row}'] = f"{reporte.fecha_inicio.strftime('%d/%m/%Y')} - {reporte.fecha_fin.strftime('%d/%m/%Y')}"
    row += 1
    
    ws[f'A{row}'] = 'Duración:'
    ws[f'A{row}'].font = info_label_font
    ws[f'B{row}'] = f"{reporte.duracion_dias} días"
    row += 1
    
    ws[f'A{row}'] = 'Creado por:'
    ws[f'A{row}'].font = info_label_font
    ws[f'B{row}'] = str(reporte.creado_por)
    row += 1
    
    ws[f'A{row}'] = 'Fecha de generación:'
    ws[f'A{row}'].font = info_label_font
    ws[f'B{row}'] = reporte.ultima_generacion.strftime('%d/%m/%Y %H:%M') if reporte.ultima_generacion else 'N/A'
    row += 2
    
    # Resumen
    if 'resumen' in datos and datos['resumen']:
        ws[f'A{row}'] = 'RESUMEN'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for key, value in datos['resumen'].items():
            ws[f'A{row}'] = key
            ws[f'A{row}'].font = info_label_font
            ws[f'B{row}'] = value
            row += 1
        row += 1
    
    # Detalles en tabla
    if 'detalles' in datos and datos['detalles']:
        ws[f'A{row}'] = 'DETALLES'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Encabezados
        headers = list(datos['detalles'][0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        row += 1
        
        # Datos
        for item in datos['detalles']:
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = item.get(header, '')
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1
    
    # Totales
    if 'totales' in datos and datos['totales']:
        ws[f'A{row}'] = 'TOTALES'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        for key, value in datos['totales'].items():
            ws[f'A{row}'] = key
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Crear la respuesta HTTP
    filename = f"reporte_{reporte.tipo}_{reporte.fecha_inicio}_{reporte.fecha_fin}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Guardar workbook en la respuesta
    wb.save(response)
    return response


def generar_pdf_reporte(reporte, datos):
    """Generar reporte en formato PDF con diseño profesional"""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab no está disponible para exportar a PDF")
    
    # Crear la respuesta HTTP
    filename = f"reporte_{reporte.tipo}_{reporte.fecha_inicio}_{reporte.fecha_fin}.pdf"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Crear PDF
    doc = SimpleDocTemplate(response, pagesize=A4, 
                          rightMargin=72, leftMargin=72,
                          topMargin=72, bottomMargin=18)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#A68932'),
        alignment=TA_CENTER,
        spaceAfter=30,
        spaceBefore=0,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
    )

    # Título principal
    title = Paragraph(f"BAR GALILEO<br/>REPORTE DE {reporte.get_tipo_display().upper()}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 20))
    
    # Información del reporte
    info_data = [
        ['Nombre:', reporte.nombre],
        ['Tipo:', reporte.get_tipo_display()],
        ['Periodo:', f"{reporte.fecha_inicio.strftime('%d/%m/%Y')} - {reporte.fecha_fin.strftime('%d/%m/%Y')}"],
        ['Duración:', f"{reporte.duracion_dias} días"],
        ['Creado por:', str(reporte.creado_por)],
        ['Fecha de generación:', reporte.ultima_generacion.strftime('%d/%m/%Y %H:%M') if reporte.ultima_generacion else 'N/A']
    ]
    
    info_table = Table(info_data, colWidths=[120, 300])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#366092')),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 25))

    # Resumen
    if 'resumen' in datos and datos['resumen']:
        subtitle = Paragraph("RESUMEN", subtitle_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 10))
        
        resumen_data = [['Concepto', 'Valor']]
        for key, value in datos['resumen'].items():
            resumen_data.append([key, str(value)])
        
        resumen_table = Table(resumen_data, colWidths=[300, 150])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5F5')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
        ]))
        
        elements.append(resumen_table)
        elements.append(Spacer(1, 20))

    # Detalles
    if 'detalles' in datos and datos['detalles']:
        subtitle = Paragraph("DETALLES", subtitle_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 10))
        
        # Preparar encabezados y datos
        headers = list(datos['detalles'][0].keys())
        detail_data = [headers]
        
        for item in datos['detalles']:
            row = [str(item.get(h, '')) for h in headers]
            detail_data.append(row)
        
        # Calcular anchos de columna
        num_cols = len(headers)
        col_width = 450 / num_cols if num_cols > 0 else 100
        col_widths = [col_width] * num_cols
        
        detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')])
        ]))
        
        elements.append(detail_table)
        elements.append(Spacer(1, 20))
    
    # Totales
    if 'totales' in datos and datos['totales']:
        subtitle = Paragraph("TOTALES", subtitle_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 10))
        
        totales_data = []
        for key, value in datos['totales'].items():
            totales_data.append([key, str(value)])
        
        totales_table = Table(totales_data, colWidths=[300, 150])
        totales_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFE6CC')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#366092')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#A68932')),
        ]))
        
        elements.append(totales_table)
    
    # Pie de página
    elements.append(Spacer(1, 30))
    footer_text = f"Generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}"
    footer = Paragraph(footer_text, ParagraphStyle('Footer', parent=styles['Normal'], 
                                                   fontSize=8, textColor=colors.grey, 
                                                   alignment=TA_CENTER))
    elements.append(footer)
    
    # Construir PDF
    doc.build(elements)
    return response


def obtener_datos_reporte_detallado(reporte):
    """
    Obtiene datos completos y detallados del reporte según su tipo
    Retorna un diccionario con 'resumen', 'detalles' y 'totales'
    """
    datos = {
        'resumen': {},
        'detalles': [],
        'totales': {}
    }
    
    try:
        if reporte.tipo == 'ventas':
            datos = obtener_datos_ventas(reporte)
        elif reporte.tipo == 'gastos':
            datos = obtener_datos_gastos(reporte)
        elif reporte.tipo == 'nominas':
            datos = obtener_datos_nominas(reporte)
        elif reporte.tipo == 'inventario':
            datos = obtener_datos_inventario(reporte)
        elif reporte.tipo == 'productos':
            datos = obtener_datos_productos(reporte)
        elif reporte.tipo == 'mesas':
            datos = obtener_datos_mesas(reporte)
        else:  # general
            datos = obtener_datos_general(reporte)
    except Exception as e:
        datos['resumen']['Error'] = str(e)
    
    return datos


def obtener_datos_ventas(reporte):
    """Obtiene datos detallados de ventas"""
    from tables.models import Factura, Pedido, PedidoItem
    
    facturas = Factura.objects.filter(
        fecha__range=[reporte.fecha_inicio, reporte.fecha_fin]
    ).select_related('pedido', 'pedido__mesa').prefetch_related('pedido__items__producto')
    
    total_ventas = facturas.aggregate(Sum('total'))['total__sum'] or Decimal('0')
    cantidad_facturas = facturas.count()
    promedio_venta = facturas.aggregate(Avg('total'))['total__avg'] or Decimal('0')
    
    # Resumen
    resumen = {
        'Total de Ventas': f"${total_ventas:,.2f}",
        'Cantidad de Facturas': cantidad_facturas,
        'Promedio por Factura': f"${promedio_venta:,.2f}",
        'Periodo': f"{reporte.fecha_inicio.strftime('%d/%m/%Y')} - {reporte.fecha_fin.strftime('%d/%m/%Y')}",
    }
    
    # Detalles
    detalles = []
    for factura in facturas:
        mesa_nombre = factura.pedido.mesa.nombre if factura.pedido.mesa else 'Sin mesa'
        items_count = factura.pedido.items.count()
        
        detalles.append({
            'Factura #': factura.numero,
            'Fecha': factura.fecha.strftime('%d/%m/%Y %H:%M'),
            'Mesa': mesa_nombre,
            'Items': items_count,
            'Total': f"${factura.total:,.2f}"
        })
    
    # Totales
    totales = {
        'TOTAL VENTAS': f"${total_ventas:,.2f}",
        'FACTURAS GENERADAS': cantidad_facturas,
    }
    
    return {
        'resumen': resumen,
        'detalles': detalles,
        'totales': totales
    }


def obtener_datos_gastos(reporte):
    """Obtiene datos detallados de gastos"""
    from expenses.models import Expense, ExpenseCategory
    
    gastos = Expense.objects.filter(
        date__range=[reporte.fecha_inicio, reporte.fecha_fin]
    ).select_related('category', 'user').order_by('-date')
    
    total_gastos = gastos.aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    cantidad_gastos = gastos.count()
    promedio_gasto = gastos.aggregate(Avg('amount'))['amount__avg'] or Decimal('0')
    
    # Gastos por categoría
    gastos_por_categoria = gastos.values('category__name').annotate(
        total=Sum('amount'),
        cantidad=Count('id')
    ).order_by('-total')
    
    # Resumen
    resumen = {
        'Total de Gastos': f"${total_gastos:,.2f}",
        'Cantidad de Gastos': cantidad_gastos,
        'Promedio por Gasto': f"${promedio_gasto:,.2f}",
        'Categorías': gastos_por_categoria.count(),
    }
    
    # Agregar categorías al resumen
    for cat in gastos_por_categoria[:5]:  # Top 5 categorías
        resumen[f"  - {cat['category__name']}"] = f"${cat['total']:,.2f} ({cat['cantidad']} gastos)"
    
    # Detalles
    detalles = []
    for gasto in gastos:
        detalles.append({
            'Fecha': gasto.date.strftime('%d/%m/%Y'),
            'Categoría': gasto.category.name,
            'Descripción': gasto.description[:50] + '...' if len(gasto.description) > 50 else gasto.description,
            'Usuario': gasto.user.username,
            'Monto': f"${gasto.amount:,.2f}"
        })
    
    # Totales
    totales = {
        'TOTAL GASTOS': f"${total_gastos:,.2f}",
        'REGISTROS': cantidad_gastos,
    }
    
    return {
        'resumen': resumen,
        'detalles': detalles,
        'totales': totales
    }


def obtener_datos_nominas(reporte):
    """Obtiene datos detallados de nóminas"""
    from nominas.models import Empleado
    
    empleados = Empleado.objects.filter(
        estado='activo'
    ).order_by('nombre')
    
    total_empleados = empleados.count()
    total_salarios = empleados.aggregate(Sum('salario'))['salario__sum'] or Decimal('0')
    promedio_salario = empleados.aggregate(Avg('salario'))['salario__avg'] or Decimal('0')
    
    # Por tipo de contrato
    por_contrato = empleados.values('tipo_contrato').annotate(
        cantidad=Count('id'),
        total_salarios=Sum('salario')
    ).order_by('-cantidad')
    
    # Resumen
    resumen = {
        'Total de Empleados Activos': total_empleados,
        'Total en Salarios': f"${total_salarios:,.2f}",
        'Promedio Salarial': f"${promedio_salario:,.2f}",
    }
    
    # Agregar por tipo de contrato
    for contrato in por_contrato:
        tipo_display = dict(Empleado.TIPO_CONTRATO_CHOICES).get(contrato['tipo_contrato'], contrato['tipo_contrato'])
        resumen[f"  - {tipo_display}"] = f"{contrato['cantidad']} empleados (${contrato['total_salarios']:,.2f})"
    
    # Detalles
    detalles = []
    for empleado in empleados:
        tipo_display = dict(Empleado.TIPO_CONTRATO_CHOICES).get(empleado.tipo_contrato, empleado.tipo_contrato)
        tiempo_servicio = (reporte.fecha_fin - empleado.fecha_contratacion).days // 365
        
        detalles.append({
            'Nombre': empleado.nombre,
            'Cargo': empleado.cargo or 'N/A',
            'Tipo Contrato': tipo_display,
            'Salario': f"${empleado.salario:,.2f}",
            'Años de Servicio': tiempo_servicio,
            'Fecha Contratación': empleado.fecha_contratacion.strftime('%d/%m/%Y')
        })
    
    # Totales
    totales = {
        'TOTAL EMPLEADOS': total_empleados,
        'TOTAL NÓMINA MENSUAL': f"${total_salarios:,.2f}",
    }
    
    return {
        'resumen': resumen,
        'detalles': detalles,
        'totales': totales
    }


def obtener_datos_inventario(reporte):
    """Obtiene datos detallados del inventario"""
    from products.models import Producto
    
    productos = Producto.objects.filter(activo=True).select_related(
        'id_categoria', 'id_proveedor', 'id_marca'
    )
    
    total_productos = productos.count()
    valor_inventario = sum(p.precio_compra * p.stock for p in productos)
    productos_bajo_stock = productos.filter(stock__lt=10).count()
    productos_sin_stock = productos.filter(stock=0).count()
    
    # Resumen
    resumen = {
        'Total de Productos': total_productos,
        'Valor del Inventario': f"${valor_inventario:,.2f}",
        'Productos con Stock Bajo (< 10)': productos_bajo_stock,
        'Productos Sin Stock': productos_sin_stock,
    }
    
    # Detalles
    detalles = []
    for producto in productos.order_by('stock')[:100]:  # Primeros 100
        categoria = producto.id_categoria.nombre_categoria if producto.id_categoria else 'Sin categoría'
        proveedor = producto.id_proveedor.nombre if producto.id_proveedor else 'Sin proveedor'
        valor_total = producto.precio_compra * producto.stock
        
        detalles.append({
            'Producto': producto.nombre,
            'Categoría': categoria,
            'Proveedor': proveedor,
            'Stock': producto.stock,
            'Precio Compra': f"${producto.precio_compra:,.2f}",
            'Precio Venta': f"${producto.precio_venta:,.2f}",
            'Valor Total': f"${valor_total:,.2f}"
        })
    
    # Totales
    totales = {
        'TOTAL PRODUCTOS': total_productos,
        'VALOR INVENTARIO': f"${valor_inventario:,.2f}",
    }
    
    return {
        'resumen': resumen,
        'detalles': detalles,
        'totales': totales
    }


def obtener_datos_productos(reporte):
    """
    Obtiene datos detallados de productos con análisis avanzados:
    - Análisis de rentabilidad (mayor/menor margen, valor potencial)
    - Alertas de stock (críticos, reorden, exceso)
    - Estadísticas por proveedor
    - Estadísticas de precios
    """
    from products.models import Producto, Categoria
    
    productos = Producto.objects.filter(activo=True).select_related(
        'id_categoria', 'id_proveedor', 'id_marca'
    )
    
    total_productos = productos.count()
    productos_activos = productos.filter(activo=True).count()
    categorias_count = Categoria.objects.all().count()
    
    # ========== ANÁLISIS DE RENTABILIDAD ==========
    productos_con_margen = []
    for p in productos:
        if p.precio_compra > 0:
            margen_porcentaje = ((p.precio_venta - p.precio_compra) / p.precio_compra * 100)
            ganancia_unitaria = p.precio_venta - p.precio_compra
            valor_potencial = ganancia_unitaria * p.stock
            
            productos_con_margen.append({
                'producto': p,
                'margen': margen_porcentaje,
                'ganancia_unitaria': ganancia_unitaria,
                'valor_potencial': valor_potencial
            })
    
    # Productos con mayor y menor margen
    productos_con_margen_ordenados = sorted(productos_con_margen, key=lambda x: x['margen'], reverse=True)
    mayor_margen = productos_con_margen_ordenados[0] if productos_con_margen_ordenados else None
    menor_margen = productos_con_margen_ordenados[-1] if productos_con_margen_ordenados else None
    
    # Productos con mayor valor potencial de venta
    productos_valor_potencial = sorted(productos_con_margen, key=lambda x: x['valor_potencial'], reverse=True)
    top_valor_potencial = productos_valor_potencial[:5] if len(productos_valor_potencial) >= 5 else productos_valor_potencial
    
    # ========== ALERTAS DE STOCK ==========
    # Stock crítico (< 5 unidades)
    productos_stock_critico = productos.filter(stock__lt=5, stock__gt=0).order_by('stock')
    # Stock bajo - necesita reorden (5-10 unidades)
    productos_reorden = productos.filter(stock__gte=5, stock__lte=10).order_by('stock')
    # Sin stock (0 unidades)
    productos_sin_stock = productos.filter(stock=0)
    # Stock excesivo (> 100 unidades)
    productos_exceso_stock = productos.filter(stock__gt=100).order_by('-stock')
    
    # ========== ESTADÍSTICAS POR PROVEEDOR ==========
    por_proveedor = productos.values('id_proveedor__nombre').annotate(
        cantidad=Count('id_producto'),
        stock_total=Sum('stock'),
        valor_compra=Sum(F('precio_compra') * F('stock')),
        valor_venta=Sum(F('precio_venta') * F('stock'))
    ).order_by('-cantidad')
    
    # ========== ESTADÍSTICAS DE PRECIOS ==========
    precios_stats = productos.aggregate(
        precio_compra_promedio=Avg('precio_compra'),
        precio_venta_promedio=Avg('precio_venta'),
        precio_compra_min=Sum('precio_compra'),
        precio_compra_max=Sum('precio_compra'),
        precio_venta_min=Sum('precio_venta'),
        precio_venta_max=Sum('precio_venta')
    )
    
    # Calcular margen promedio
    if productos_con_margen:
        margen_promedio = sum(p['margen'] for p in productos_con_margen) / len(productos_con_margen)
    else:
        margen_promedio = 0
    
    # Valor total del inventario
    valor_inventario_compra = sum(p.precio_compra * p.stock for p in productos)
    valor_inventario_venta = sum(p.precio_venta * p.stock for p in productos)
    ganancia_potencial_total = valor_inventario_venta - valor_inventario_compra
    
    # Por categoría
    por_categoria = productos.values('id_categoria__nombre_categoria').annotate(
        cantidad=Count('id_producto'),
        stock_total=Sum('stock')
    ).order_by('-cantidad')
    
    # ========== RESUMEN MEJORADO ==========
    resumen = {
        '=== INFORMACIÓN GENERAL ===': '',
        'Total de Productos': total_productos,
        'Productos Activos': productos_activos,
        'Categorías': categorias_count,
        'Proveedores': por_proveedor.count(),
    }
    
    # Separador
    resumen['─' * 50] = ''
    
    # Análisis de rentabilidad
    resumen['=== ANÁLISIS DE RENTABILIDAD ==='] = ''
    resumen['Margen Promedio'] = f"{margen_promedio:.2f}%"
    resumen['Valor Inventario (Compra)'] = f"${valor_inventario_compra:,.2f}"
    resumen['Valor Inventario (Venta)'] = f"${valor_inventario_venta:,.2f}"
    resumen['Ganancia Potencial Total'] = f"${ganancia_potencial_total:,.2f}"
    
    # Agregar productos con mayor y menor margen
    if mayor_margen:
        resumen['Producto con Mayor Margen'] = f"{mayor_margen['producto'].nombre} ({mayor_margen['margen']:.1f}%)"
    if menor_margen:
        resumen['Producto con Menor Margen'] = f"{menor_margen['producto'].nombre} ({menor_margen['margen']:.1f}%)"
    
    # Separador
    resumen['─' * 50 + ' '] = ''
    
    # Agregar top productos por valor potencial
    resumen['=== TOP 5 VALOR POTENCIAL ==='] = ''
    for i, item in enumerate(top_valor_potencial, 1):
        resumen[f"  {i}. {item['producto'].nombre}"] = f"${item['valor_potencial']:,.2f}"
    
    # Separador
    resumen['─' * 50 + '  '] = ''
    
    # Alertas de stock
    resumen['=== ALERTAS DE STOCK ==='] = ''
    resumen['🔴 Stock Crítico (< 5)'] = productos_stock_critico.count()
    resumen['🟡 Requiere Reorden (5-10)'] = productos_reorden.count()
    resumen['⚫ Sin Stock'] = productos_sin_stock.count()
    resumen['🔵 Stock Excesivo (> 100)'] = productos_exceso_stock.count()
    
    # Separador
    resumen['─' * 50 + '   '] = ''
    
    # Estadísticas por proveedor
    resumen['=== TOP 5 PROVEEDORES ==='] = ''
    for prov in por_proveedor[:5]:
        prov_nombre = prov['id_proveedor__nombre'] or 'Sin proveedor'
        valor = prov['valor_compra'] or 0
        resumen[f"  - {prov_nombre}"] = f"{prov['cantidad']} productos (${valor:,.2f})"
    
    # Separador
    resumen['─' * 50 + '    '] = ''
    
    # Estadísticas de precios
    resumen['=== ESTADÍSTICAS DE PRECIOS ==='] = ''
    resumen['Precio Compra Promedio'] = f"${precios_stats['precio_compra_promedio'] or 0:,.2f}"
    resumen['Precio Venta Promedio'] = f"${precios_stats['precio_venta_promedio'] or 0:,.2f}"
    
    # Separador
    resumen['─' * 50 + '     '] = ''
    
    # Top 5 categorías
    resumen['=== TOP 5 CATEGORÍAS ==='] = ''
    for cat in por_categoria[:5]:
        cat_nombre = cat['id_categoria__nombre_categoria'] or 'Sin categoría'
        resumen[f"  - {cat_nombre}"] = f"{cat['cantidad']} productos (Stock: {cat['stock_total'] or 0})"
    
    # ========== DETALLES MEJORADOS ==========
    detalles = []
    
    # Incluir todos los productos con información completa
    for producto in productos.order_by('-stock')[:200]:  # Primeros 200
        categoria = producto.id_categoria.nombre_categoria if producto.id_categoria else 'Sin categoría'
        marca = producto.id_marca.marca if producto.id_marca else 'Sin marca'
        proveedor = producto.id_proveedor.nombre if producto.id_proveedor else 'Sin proveedor'
        
        if producto.precio_compra > 0:
            margen = ((producto.precio_venta - producto.precio_compra) / producto.precio_compra * 100)
            ganancia_unitaria = producto.precio_venta - producto.precio_compra
            valor_potencial = ganancia_unitaria * producto.stock
        else:
            margen = 0
            ganancia_unitaria = 0
            valor_potencial = 0
        
        # Determinar alerta de stock
        if producto.stock == 0:
            alerta_stock = '⚫ Sin Stock'
        elif producto.stock < 5:
            alerta_stock = '🔴 Crítico'
        elif producto.stock <= 10:
            alerta_stock = '🟡 Reorden'
        elif producto.stock > 100:
            alerta_stock = '🔵 Exceso'
        else:
            alerta_stock = '✅ Normal'
        
        detalles.append({
            'Producto': producto.nombre,
            'Categoría': categoria,
            'Marca': marca,
            'Proveedor': proveedor,
            'Stock': producto.stock,
            'Alerta': alerta_stock,
            'Precio Compra': f"${producto.precio_compra:,.2f}",
            'Precio Venta': f"${producto.precio_venta:,.2f}",
            'Margen': f"{margen:.1f}%",
            'Ganancia Unit.': f"${ganancia_unitaria:,.2f}",
            'Valor Potencial': f"${valor_potencial:,.2f}"
        })
    
    # ========== TOTALES MEJORADOS ==========
    totales = {
        '=== INVENTARIO ===': '',
        'TOTAL PRODUCTOS': total_productos,
        'PRODUCTOS ACTIVOS': productos_activos,
        'VALOR INVENTARIO (COMPRA)': f"${valor_inventario_compra:,.2f}",
        'VALOR INVENTARIO (VENTA)': f"${valor_inventario_venta:,.2f}",
    }
    
    # Separador
    totales['═' * 50] = ''
    
    # Rentabilidad
    totales['=== RENTABILIDAD ==='] = ''
    totales['GANANCIA POTENCIAL TOTAL'] = f"${ganancia_potencial_total:,.2f}"
    totales['MARGEN PROMEDIO'] = f"{margen_promedio:.2f}%"
    
    # Separador
    totales['═' * 50 + ' '] = ''
    
    # Alertas
    totales['=== ALERTAS ==='] = ''
    totales['🔴 STOCK CRÍTICO'] = productos_stock_critico.count()
    totales['🟡 REQUIERE REORDEN'] = productos_reorden.count()
    totales['⚫ SIN STOCK'] = productos_sin_stock.count()
    totales['🔵 STOCK EXCESIVO'] = productos_exceso_stock.count()
    
    return {
        'resumen': resumen,
        'detalles': detalles,
        'totales': totales
    }


def obtener_datos_mesas(reporte):
    """Obtiene datos detallados de mesas y pedidos"""
    from tables.models import Mesa, Pedido, Factura
    
    pedidos = Pedido.objects.filter(
        fecha_creacion__range=[reporte.fecha_inicio, reporte.fecha_fin]
    ).select_related('mesa').prefetch_related('items')
    
    mesas = Mesa.objects.all()
    facturas = Factura.objects.filter(
        fecha__range=[reporte.fecha_inicio, reporte.fecha_fin]
    )
    
    total_mesas = mesas.count()
    total_pedidos = pedidos.count()
    total_facturado = facturas.aggregate(Sum('total'))['total__sum'] or Decimal('0')
    pedidos_facturados = pedidos.filter(estado='facturado').count()
    pedidos_cancelados = pedidos.filter(estado='cancelado').count()
    
    # Resumen
    resumen = {
        'Total de Mesas': total_mesas,
        'Total de Pedidos': total_pedidos,
        'Pedidos Facturados': pedidos_facturados,
        'Pedidos Cancelados': pedidos_cancelados,
        'Total Facturado': f"${total_facturado:,.2f}",
        'Promedio por Pedido': f"${(total_facturado / pedidos_facturados if pedidos_facturados > 0 else 0):,.2f}",
    }
    
    # Detalles por pedido
    detalles = []
    for pedido in pedidos.order_by('-fecha_creacion')[:100]:  # Últimos 100
        mesa_nombre = pedido.mesa.nombre if pedido.mesa else 'Sin mesa'
        items_count = pedido.items.count()
        total_pedido = pedido.total()
        
        detalles.append({
            'Pedido #': pedido.id,
            'Mesa': mesa_nombre,
            'Fecha': pedido.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            'Estado': pedido.get_estado_display(),
            'Items': items_count,
            'Total': f"${total_pedido:,.2f}"
        })
    
    # Totales
    totales = {
        'TOTAL PEDIDOS': total_pedidos,
        'TOTAL FACTURADO': f"${total_facturado:,.2f}",
    }
    
    return {
        'resumen': resumen,
        'detalles': detalles,
        'totales': totales
    }


def obtener_datos_general(reporte):
    """Obtiene datos generales del sistema"""
    from tables.models import Factura, Pedido
    from expenses.models import Expense
    from products.models import Producto
    from nominas.models import Empleado
    
    # Obtener datos de todos los módulos
    facturas = Factura.objects.filter(fecha__range=[reporte.fecha_inicio, reporte.fecha_fin])
    gastos = Expense.objects.filter(date__range=[reporte.fecha_inicio, reporte.fecha_fin])
    
    total_ventas = facturas.aggregate(Sum('total'))['total__sum'] or Decimal('0')
    total_gastos = gastos.aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    utilidad_bruta = total_ventas - total_gastos
    
    # Resumen
    resumen = {
        'Periodo': f"{reporte.fecha_inicio.strftime('%d/%m/%Y')} - {reporte.fecha_fin.strftime('%d/%m/%Y')}",
        'Duración': f"{reporte.duracion_dias} días",
        'Total Ventas': f"${total_ventas:,.2f}",
        'Total Gastos': f"${total_gastos:,.2f}",
        'Utilidad Bruta': f"${utilidad_bruta:,.2f}",
        'Margen': f"{(utilidad_bruta / total_ventas * 100 if total_ventas > 0 else 0):.2f}%",
        'Facturas Generadas': facturas.count(),
        'Gastos Registrados': gastos.count(),
        'Productos Activos': Producto.objects.filter(activo=True).count(),
        'Empleados Activos': Empleado.objects.filter(estado='activo').count(),
    }
    
    # Detalles - Resumen por día
    detalles = []
    from datetime import timedelta
    current_date = reporte.fecha_inicio
    while current_date <= reporte.fecha_fin:
        ventas_dia = facturas.filter(fecha__date=current_date).aggregate(Sum('total'))['total__sum'] or Decimal('0')
        gastos_dia = gastos.filter(date=current_date).aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
        utilidad_dia = ventas_dia - gastos_dia
        
        if ventas_dia > 0 or gastos_dia > 0:  # Solo mostrar días con actividad
            detalles.append({
                'Fecha': current_date.strftime('%d/%m/%Y'),
                'Ventas': f"${ventas_dia:,.2f}",
                'Gastos': f"${gastos_dia:,.2f}",
                'Utilidad': f"${utilidad_dia:,.2f}"
            })
        
        current_date += timedelta(days=1)
    
    # Totales
    totales = {
        'TOTAL VENTAS': f"${total_ventas:,.2f}",
        'TOTAL GASTOS': f"${total_gastos:,.2f}",
        'UTILIDAD BRUTA': f"${utilidad_bruta:,.2f}",
    }
    
    return {
        'resumen': resumen,
        'detalles': detalles,
        'totales': totales
    }
