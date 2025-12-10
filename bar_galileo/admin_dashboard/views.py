from django.shortcuts import render
from django.views.generic import TemplateView
from products.models import Producto, Categoria
from tables.models import Mesa, Pedido, PedidoItem, Factura
from expenses.models import Expense
from nominas.models import Empleado, Pago, Bonificacion
from roles.decorators import permission_required
from django.utils.decorators import method_decorator
from notifications.utils import notificar_usuario
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncDay
from datetime import datetime, date, timedelta
from django.http import HttpResponse
import csv
from io import BytesIO
from django.template.loader import render_to_string
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

def get_date_range(period):
    today = datetime.now().date()
    if period == 'month':
        start_date = today.replace(day=1)
        # Go to the next month and subtract one day to get the end of the current month
        next_month = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)
        end_date = next_month - timedelta(days=1)
    elif period == 'quarter':
        current_quarter = (today.month - 1) // 3 + 1
        start_month = 3 * current_quarter - 2
        start_date = today.replace(month=start_month, day=1)
        end_month = 3 * current_quarter
        # handle end of year
        if end_month == 12:
            end_date = today.replace(year=today.year, month=12, day=31)
        else:
            end_date = (today.replace(month=end_month + 1, day=1)) - timedelta(days=1)
    else: # all
        start_date = None
        end_date = None
    return start_date, end_date

@method_decorator(permission_required('dashboard', 'ver'), name='dispatch')
class DashboardView(TemplateView):
    template_name = 'dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        period = self.request.GET.get('period', 'month') # Default to month
        start_date, end_date = get_date_range(period)

        # Filtered querysets
        facturas_periodo = Factura.objects.all()
        pedidos_periodo = PedidoItem.objects.filter(pedido__factura__isnull=False)
        gastos_periodo = Expense.objects.all()
        pagos_periodo = Pago.objects.all()

        if start_date and end_date:
            facturas_periodo = facturas_periodo.filter(fecha__date__range=[start_date, end_date])
            pedidos_periodo = pedidos_periodo.filter(pedido__factura__fecha__date__range=[start_date, end_date])
            gastos_periodo = gastos_periodo.filter(date__range=[start_date, end_date])
            pagos_periodo = pagos_periodo.filter(fecha_pago__range=[start_date, end_date])

        context['productos'] = Producto.objects.count()
        context['categorias'] = Categoria.objects.count()
        context['mesas'] = Mesa.objects.count()

        # Data for charts
        context['total_mesas_disponibles'] = Mesa.objects.filter(estado='disponible').count()
        context['total_mesas_ocupadas'] = Mesa.objects.filter(estado='ocupada').count()
        context['total_mesas_reservadas'] = Mesa.objects.filter(estado='reservada').count()
        context['total_mesas_fuera_de_servicio'] = Mesa.objects.filter(estado='fuera de servicio').count()

        context['top_5_productos_mas_vendidos'] = pedidos_periodo.values('producto__nombre').annotate(total_vendido=Sum('cantidad')).order_by('-total_vendido')[:5]

        # New data
        context['ingresos_totales'] = facturas_periodo.aggregate(total=Sum('total'))['total'] or 0
        context['productos_mas_stock'] = Producto.objects.order_by('-stock')[:5]
        context['valor_total_stock'] = Producto.objects.aggregate(total=Sum(F('stock') * F('precio_compra')))['total'] or 0

        # Monthly sales (or period sales)
        ventas_periodo = facturas_periodo.annotate(dia=TruncDay('fecha')).values('dia').annotate(total_dia=Sum('total')).order_by('dia')
        context['ventas_periodo'] = list(ventas_periodo) # Convert to list to be able to serialize it

        # Total profit
        ganancia_total = pedidos_periodo.annotate(
            ganancia_item=ExpressionWrapper(
                (F('precio_unitario') - F('producto__precio_compra')) * F('cantidad'),
                output_field=DecimalField()
            )
        ).aggregate(total=Sum('ganancia_item'))['total'] or 0
        context['ganancia_total'] = ganancia_total

        context['gastos_totales'] = gastos_periodo.aggregate(total=Sum('amount'))['total'] or 0

        # Nóminas data
        context['total_empleados'] = Empleado.objects.filter(estado='activo').count()
        context['total_empleados_inactivos'] = Empleado.objects.filter(estado='inactivo').count()
        context['nominas_pagadas'] = pagos_periodo.aggregate(total=Sum('monto'))['total'] or 0
        context['bonificaciones_activas'] = Bonificacion.objects.filter(activa=True).count()

        # Total de bonificaciones pagadas como pagos de tipo "bono" en el periodo
        bonificaciones_periodo = pagos_periodo.filter(tipo='bono').aggregate(total_bonificaciones=Sum('monto'))['total_bonificaciones'] or 0
        context['bonificaciones_pagadas'] = bonificaciones_periodo

        # Pagos por empleado (top 5 con mayores pagos en el periodo)
        top_pagos = pagos_periodo.values('empleado__nombre').annotate(
            total_pagado=Sum('monto')
        ).order_by('-total_pagado')[:5]
        context['top_pagos_empleados'] = top_pagos

        context['selected_period'] = period

        # Enviar notificación al usuario actual
        # if self.request.user.is_authenticated:
        #     mensaje = f"¡Bienvenido al dashboard, {self.request.user.first_name or self.request.user.username}!"
        #     notificar_usuario(self.request.user, mensaje)
        return context


def _build_dashboard_context(request):
    """Helper para construir el mismo contexto que DashboardView (usado por export)."""
    period = request.GET.get('period', 'month')
    start_date, end_date = get_date_range(period)

    facturas_periodo = Factura.objects.all()
    pedidos_periodo = PedidoItem.objects.filter(pedido__factura__isnull=False)
    gastos_periodo = Expense.objects.all()

    if start_date and end_date:
        facturas_periodo = facturas_periodo.filter(fecha__date__range=[start_date, end_date])
        pedidos_periodo = pedidos_periodo.filter(pedido__factura__fecha__date__range=[start_date, end_date])
        gastos_periodo = gastos_periodo.filter(date__range=[start_date, end_date])

    ingresos_totales = facturas_periodo.aggregate(total=Sum('total'))['total'] or 0
    ventas_periodo = facturas_periodo.annotate(dia=TruncDay('fecha')).values('dia').annotate(total_dia=Sum('total')).order_by('dia')
    ganancia_total = pedidos_periodo.annotate(
        ganancia_item=ExpressionWrapper(
            (F('precio_unitario') - F('producto__precio_compra')) * F('cantidad'),
            output_field=DecimalField()
        )
    ).aggregate(total=Sum('ganancia_item'))['total'] or 0
    valor_total_stock = Producto.objects.aggregate(total=Sum(F('stock') * F('precio_compra')))['total'] or 0
    gastos_totales = gastos_periodo.aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'period': period,
        'ingresos_totales': ingresos_totales,
        'ventas_periodo': list(ventas_periodo),
        'ganancia_total': ganancia_total,
        'valor_total_stock': valor_total_stock,
        'gastos_totales': gastos_totales,
    }
    return context


def export_dashboard(request, fmt):
    """Exportar estadísticas generales en CSV o PDF (fallback HTML si no hay reportlab).
    fmt: 'csv' o 'pdf'
    """
    try:
        ctx = _build_dashboard_context(request)
    except Exception as e:
        return HttpResponse(f"Error generando reporte: {e}", status=500)

    if fmt == 'csv' or fmt == 'xlsx':
        # If XLSX requested and openpyxl is available, generate Excel file
        if fmt == 'xlsx' and openpyxl is not None:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Estadisticas'

            # Encabezados y métricas
            ws.append(['Métrica', 'Valor'])
            ws.append(['Ingresos Totales', float(ctx['ingresos_totales'])])
            ws.append(['Ganancia Total', float(ctx['ganancia_total'])])
            ws.append(['Valor Total del Stock', float(ctx['valor_total_stock'])])
            ws.append(['Gastos Totales', float(ctx['gastos_totales'])])
            ws.append([])
            ws.append(['Ventas por Día'])
            ws.append(['Fecha', 'Total'])
            for v in ctx['ventas_periodo']:
                    fecha = v.get('dia')
                    total = v.get('total_dia')
                    # Keep the date/datetime object so Excel recognizes it as a date
                    if hasattr(fecha, 'isoformat'):
                        fecha_val = fecha
                    else:
                        try:
                            fecha_val = datetime.fromisoformat(str(fecha))
                        except Exception:
                            fecha_val = None
                    ws.append([fecha_val, float(total or 0)])

            # Ajustar anchos de columnas
            for i, col in enumerate(ws.columns, 1):
                max_length = 0
                for cell in col:
                    try:
                        value = str(cell.value)
                    except Exception:
                        value = ''
                    if value and len(value) > max_length:
                        max_length = len(value)
                ws.column_dimensions[get_column_letter(i)].width = min(max_length + 2, 50)

            # Apply number/date formats: detect date cells and totals
            for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
                cell_date = row[0]
                cell_total = row[1]
                try:
                    if isinstance(cell_date.value, (datetime, date)):
                        cell_date.number_format = 'dd/mm/yyyy'
                except Exception:
                    pass
                try:
                    cell_total.number_format = '#,##0.00'
                except Exception:
                    pass

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="estadisticas_generales.xlsx"'
            return response

        # Fallback: generate CSV with formatted values
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="estadisticas_generales.csv"'
        writer = csv.writer(response)

        # Escribir métricas principales
        writer.writerow(['Métrica', 'Valor'])
        writer.writerow(['Ingresos Totales', f"{float(ctx['ingresos_totales']):.2f}"])
        writer.writerow(['Ganancia Total', f"{float(ctx['ganancia_total']):.2f}"])
        writer.writerow(['Valor Total del Stock', f"{float(ctx['valor_total_stock']):.2f}"])
        writer.writerow(['Gastos Totales', f"{float(ctx['gastos_totales']):.2f}"])
        writer.writerow([])
        writer.writerow(['Ventas por Día'])
        writer.writerow(['Fecha', 'Total'])
        for v in ctx['ventas_periodo']:
            fecha = v.get('dia')
            total = v.get('total_dia')
            fecha_str = fecha.isoformat() if hasattr(fecha, 'isoformat') else str(fecha)
            total_str = f"{float(total or 0):.2f}"
            writer.writerow([fecha_str, total_str])

        return response

    elif fmt == 'pdf':
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
        except Exception:
            html = render_to_string('admin_dashboard/report_pdf.html', {
                'estadisticas': ctx
            })
            return HttpResponse(html, content_type='text/html')

        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        x = 40
        y = height - 40

        p.setFont('Helvetica-Bold', 16)
        p.drawString(x, y, 'Estadísticas Generales')
        y -= 30

        p.setFont('Helvetica', 12)
        p.drawString(x, y, f"Ingresos Totales: {ctx['ingresos_totales']}")
        y -= 18
        p.drawString(x, y, f"Ganancia Total: {ctx['ganancia_total']}")
        y -= 18
        p.drawString(x, y, f"Valor Total del Stock: {ctx['valor_total_stock']}")
        y -= 18
        p.drawString(x, y, f"Gastos Totales: {ctx['gastos_totales']}")
        y -= 24

        p.drawString(x, y, 'Ventas por Día:')
        y -= 18
        for v in ctx['ventas_periodo']:
            if y < 60:
                p.showPage()
                y = height - 40
            fecha = v.get('dia')
            total = v.get('total_dia')
            p.drawString(x, y, f"{fecha}: {total}")
            y -= 14

        p.showPage()
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="estadisticas_generales.pdf"'
        response.write(pdf)
        return response

    else:
        return HttpResponse('Formato no soportado', status=400)
