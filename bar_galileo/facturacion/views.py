from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q
from django.utils import timezone
from datetime import datetime
from django.core.paginator import Paginator
from tables.models import Factura
from .models import FacturacionManager
from roles.decorators import permission_required
from django.contrib.auth.decorators import login_required
from decimal import InvalidOperation
import logging

logger = logging.getLogger(__name__)

@login_required
@permission_required('facturacion', 'ver')
def lista_facturas(request):
    """
    Vista para mostrar la lista de facturas con filtros de búsqueda
    """
    # Obtener parámetros de búsqueda
    busqueda = request.GET.get('busqueda', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Convertir fechas si existen
    fecha_inicio_obj = None
    fecha_fin_obj = None
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            pass
    
    try:
        # Obtener facturas usando el nuevo manager
        facturas = FacturacionManager.obtener_facturas_con_filtros(
            busqueda=busqueda,
            fecha_inicio=fecha_inicio_obj,
            fecha_fin=fecha_fin_obj
        )
        
        # Paginación
        paginator = Paginator(facturas, 10)  # 10 facturas por página
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # Estadísticas
        estadisticas = FacturacionManager.obtener_estadisticas()
        
    except Exception as e:
        logger.error(f"Error en lista_facturas: {e}")
        messages.error(request, f"Error cargando facturas: {e}")
        page_obj = Paginator([], 10).get_page(1)
        estadisticas = {
            'total_facturas': 0,
            'total_ingresos': 0,
            'facturas_hoy': 0,
            'ingresos_hoy': 0,
        }
    
    context = {
        'page_obj': page_obj,
        'busqueda': busqueda,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'estadisticas': estadisticas,
    }
    
    return render(request, 'facturacion/lista_facturas.html', context)

@login_required
@permission_required('facturacion', 'ver')
def detalle_factura(request, factura_id):
    """
    Vista para mostrar el detalle de una factura específica
    """
    try:
        factura = FacturacionManager.obtener_factura_por_id(factura_id)
        
        # Si el método normal falla, intentamos con el método seguro
        if not factura:
            factura = FacturacionManager.obtener_factura_por_id_seguro(factura_id)
            
            if factura:
                messages.warning(request, 'Esta factura contiene algunos datos corruptos. Se muestra información limitada.')
            else:
                messages.error(request, 'Factura no encontrada.')
                return redirect('facturacion:lista_facturas')
        
        context = {
            'factura': factura,
        }
        
        return render(request, 'facturacion/detalle_factura.html', context)
        
    except InvalidOperation as e:
        logger.error(f"Error de datos corruptos en factura {factura_id}: {e}")
        # Intentar con el método seguro
        try:
            factura = FacturacionManager.obtener_factura_por_id_seguro(factura_id)
            if factura:
                messages.warning(request, 'Esta factura contiene datos corruptos. Se muestra información limitada.')
                context = {'factura': factura}
                return render(request, 'facturacion/detalle_factura.html', context)
            else:
                messages.error(request, 'Factura no encontrada.')
                return redirect('facturacion:lista_facturas')
        except Exception:
            messages.error(request, 'La factura contiene datos corruptos y no puede ser mostrada.')
            return redirect('facturacion:lista_facturas')
    except Exception as e:
        logger.error(f"Error inesperado en detalle_factura para ID {factura_id}: {e}")
        messages.error(request, f'Error al cargar la factura: {str(e)}')
        return redirect('facturacion:lista_facturas')

@login_required
@permission_required('facturacion', 'eliminar')
def eliminar_factura(request, factura_id):
    """
    Vista para eliminar una factura
    """
    factura = None
    try:
        # Intentar obtener la factura con el método normal
        factura = FacturacionManager.obtener_factura_por_id(factura_id)
        
        # Si falla, intentar con el método seguro
        if not factura:
            factura = FacturacionManager.obtener_factura_por_id_seguro(factura_id)
            if factura:
                messages.warning(request, 'Esta factura contiene datos corruptos pero se puede eliminar.')
        
        if not factura:
            messages.error(request, 'Factura no encontrada.')
            return redirect('facturacion:lista_facturas')
            
    except InvalidOperation as e:
        logger.error(f"Error de datos corruptos al cargar factura {factura_id} para eliminar: {e}")
        # Intentar con el método seguro
        try:
            factura = FacturacionManager.obtener_factura_por_id_seguro(factura_id)
            if factura:
                messages.warning(request, 'Esta factura contiene datos corruptos pero se puede eliminar.')
            else:
                messages.error(request, 'Factura no encontrada.')
                return redirect('facturacion:lista_facturas')
        except Exception:
            messages.error(request, 'Error: La factura contiene datos corruptos y no puede ser procesada.')
            return redirect('facturacion:lista_facturas')
    except Exception as e:
        logger.error(f"Error inesperado al cargar factura {factura_id} para eliminar: {e}")
        messages.error(request, f'Error al cargar la factura: {str(e)}')
        return redirect('facturacion:lista_facturas')
    
    if request.method == 'POST':
        try:
            # Eliminar usando SQL directo para evitar problemas con datos corruptos
            from django.db import connection
            with connection.cursor() as cursor:
                # Obtener el número antes de eliminar
                cursor.execute("SELECT numero FROM tables_factura WHERE id = %s", [factura_id])
                numero_result = cursor.fetchone()
                numero_factura = numero_result[0] if numero_result else f"ID-{factura_id}"
                
                # Eliminar la factura
                cursor.execute("DELETE FROM tables_factura WHERE id = %s", [factura_id])
                
            messages.success(request, f'Factura #{numero_factura} eliminada exitosamente.')
            return redirect('facturacion:lista_facturas')
            
        except Exception as e:
            logger.error(f"Error al eliminar factura {factura_id}: {e}")
            messages.error(request, f'Error al eliminar la factura: {str(e)}')
            return redirect('facturacion:lista_facturas')
    
    context = {
        'factura': factura,
    }
    
    return render(request, 'facturacion/confirmar_eliminar.html', context)

@login_required
@permission_required('facturacion', 'ver')
def buscar_facturas_ajax(request):
    """
    Vista AJAX para búsqueda de facturas en tiempo real
    """
    if request.method == 'GET':
        busqueda = request.GET.get('busqueda', '')
        
        facturas = FacturacionManager.obtener_facturas_con_filtros(busqueda=busqueda)[:10]
        
        resultados = []
        for factura in facturas:
            resultados.append({
                'id': factura.id,
                'numero': factura.numero,
                'fecha': factura.fecha.strftime('%d/%m/%Y %H:%M'),
                'mesa': factura.pedido.mesa.nombre if factura.pedido.mesa else 'Sin mesa',
                'total': str(factura.total),
            })
        
        return JsonResponse({'facturas': resultados})
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
@permission_required('facturacion', 'administrar')
def diagnostico_facturas(request):
    """
    Vista para diagnosticar facturas con datos corruptos
    """
    try:
        facturas_corruptas = FacturacionManager.verificar_facturas_corruptas()
        
        context = {
            'facturas_corruptas': facturas_corruptas,
            'total_corruptas': len(facturas_corruptas),
        }
        
        return render(request, 'facturacion/diagnostico_facturas.html', context)
        
    except Exception as e:
        logger.error(f"Error en diagnostico_facturas: {e}")
        messages.error(request, f'Error al ejecutar diagnóstico: {str(e)}')
        return redirect('facturacion:lista_facturas')

# ===== FUNCIONES DE EXPORTACIÓN =====

import io
import csv
from django.http import HttpResponse
from django.template.loader import get_template
from django.conf import settings
import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

@login_required
@permission_required('facturacion', 'ver')
def exportar_facturas_csv(request):
    """Exportar facturas a CSV"""
    # Obtener parámetros de filtro
    busqueda = request.GET.get('busqueda', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Convertir fechas si existen
    fecha_inicio_obj = None
    fecha_fin_obj = None
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            pass

    # Obtener facturas con filtros
    facturas = FacturacionManager.obtener_facturas_con_filtros(
        busqueda=busqueda,
        fecha_inicio=fecha_inicio_obj,
        fecha_fin=fecha_fin_obj
    )

    # Crear la respuesta HTTP
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="facturas.csv"'

    writer = csv.writer(response)
    
    # Escribir encabezados
    writer.writerow([
        'Número',
        'Fecha',
        'Mesa',
        'Total',
        'Items',
        'Estado'
    ])
    
    # Escribir datos
    for factura in facturas:
        try:
            writer.writerow([
                factura.numero,
                factura.fecha.strftime('%d/%m/%Y %H:%M'),
                factura.pedido.mesa.nombre if factura.pedido.mesa else 'Sin mesa',
                f'${factura.total:,.2f}',
                len(factura.pedido.items.all()) if factura.pedido else 0,
                'Pagada' if factura.fecha else 'Pendiente'
            ])
        except Exception as e:
            logger.error(f"Error exportando factura {factura.id}: {e}")
            continue

    return response

@login_required
@permission_required('facturacion', 'ver')
def exportar_facturas_xlsx(request):
    """Exportar facturas a Excel (XLSX)"""
    if not OPENPYXL_AVAILABLE:
        messages.error(request, 'La exportación a Excel no está disponible. Instale openpyxl.')
        return redirect('facturacion:lista_facturas')
    
    # Obtener parámetros de filtro
    busqueda = request.GET.get('busqueda', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Convertir fechas si existen
    fecha_inicio_obj = None
    fecha_fin_obj = None
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            pass

    # Obtener facturas con filtros
    facturas = FacturacionManager.obtener_facturas_con_filtros(
        busqueda=busqueda,
        fecha_inicio=fecha_inicio_obj,
        fecha_fin=fecha_fin_obj
    )

    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = ['Número', 'Fecha', 'Mesa', 'Total', 'Items', 'Estado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Datos
    for row, factura in enumerate(facturas, 2):
        try:
            ws.cell(row=row, column=1, value=factura.numero)
            ws.cell(row=row, column=2, value=factura.fecha.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=3, value=factura.pedido.mesa.nombre if factura.pedido.mesa else 'Sin mesa')
            ws.cell(row=row, column=4, value=float(factura.total))
            ws.cell(row=row, column=5, value=len(factura.pedido.items.all()) if factura.pedido else 0)
            ws.cell(row=row, column=6, value='Pagada' if factura.fecha else 'Pendiente')
        except Exception as e:
            logger.error(f"Error exportando factura {factura.id}: {e}")
            continue

    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Crear la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="facturas.xlsx"'

    # Guardar workbook en la respuesta
    wb.save(response)
    return response

@login_required
@permission_required('facturacion', 'ver')
def exportar_facturas_pdf(request):
    """Exportar facturas a PDF"""
    if not REPORTLAB_AVAILABLE:
        messages.error(request, 'La exportación a PDF no está disponible. Instale reportlab.')
        return redirect('facturacion:lista_facturas')
    
    # Obtener parámetros de filtro
    busqueda = request.GET.get('busqueda', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    
    # Convertir fechas si existen
    fecha_inicio_obj = None
    fecha_fin_obj = None
    
    if fecha_inicio:
        try:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        except ValueError:
            pass

    # Obtener facturas con filtros
    facturas = FacturacionManager.obtener_facturas_con_filtros(
        busqueda=busqueda,
        fecha_inicio=fecha_inicio_obj,
        fecha_fin=fecha_fin_obj
    )

    # Crear la respuesta HTTP
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="facturas.pdf"'

    # Crear PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        alignment=1,  # Centrado
        spaceAfter=30,
    )

    # Título
    title = Paragraph("Reporte de Facturas - Bar Galileo", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))

    # Preparar datos para la tabla
    data = [['Número', 'Fecha', 'Mesa', 'Total', 'Items']]
    
    for factura in facturas:
        try:
            data.append([
                factura.numero,
                factura.fecha.strftime('%d/%m/%Y'),
                factura.pedido.mesa.nombre if factura.pedido.mesa else 'Sin mesa',
                f'${factura.total:,.2f}',
                str(len(factura.pedido.items.all()) if factura.pedido else 0)
            ])
        except Exception as e:
            logger.error(f"Error exportando factura {factura.id}: {e}")
            continue

    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    return response
